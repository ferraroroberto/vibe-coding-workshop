# Simple script to test basic imports and functionality of workshop libraries
# Testing libraries from requirements.txt in root folder
import logging
import os
import sys

log = logging.getLogger(__name__)


def check(msg, success):
    # Green check for success, red cross for failure
    if success:
        log.info("\033[92m✔\033[0m %s", msg)
    else:
        log.info("\033[91m✘\033[0m %s", msg)


def main():
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    log.info("=" * 55)
    log.info("  Workshop Library Test Suite")
    log.info("=" * 55)
    log.info("  Python: %s", sys.version.split()[0])
    log.info("  CWD:    %s", os.getcwd())
    log.info("=" * 55)

    # --- Faker ---
    log.info("--- Faker ---")
    try:
        import faker
        fake = faker.Faker()
        version = getattr(faker, '__version__', 'installed')
        log.info("  Version: %s", version)
        log.info("  Sample name: %s", fake.name())
        check("Faker test", True)
    except Exception as e:
        check(f"Faker test failed: {e}", False)

    # --- DuckDB ---
    log.info("--- DuckDB ---")
    try:
        import duckdb
        conn = duckdb.connect()
        result = conn.execute("SELECT 1 as test").fetchone()
        log.info("  Version: %s", duckdb.__version__)
        log.info("  Test query result: %s", result)
        conn.close()
        check("DuckDB test", True)
    except Exception as e:
        check(f"DuckDB test failed: {e}", False)

    # --- Pandas ---
    log.info("--- Pandas ---")
    try:
        import pandas as pd
        log.info("  Version: %s", pd.__version__)
        df = pd.DataFrame({'A': [1, 2], 'B': [3, 4]})
        assert len(df) == 2
        log.info("  DataFrame created: %s", df.shape)
        check("Pandas test", True)
    except Exception as e:
        check(f"Pandas test failed: {e}", False)

    # --- NumPy ---
    log.info("--- NumPy ---")
    try:
        import numpy as np
        log.info("  Version: %s", np.__version__)
        arr = np.array([1, 2, 3, 4, 5])
        log.info("  Array mean: %s", arr.mean())
        check("NumPy test", True)
    except Exception as e:
        check(f"NumPy test failed: {e}", False)

    # --- Matplotlib ---
    log.info("--- Matplotlib ---")
    try:
        import matplotlib
        matplotlib.use("Agg")  # Non-interactive backend
        import matplotlib.pyplot as plt
        log.info("  Version: %s", matplotlib.__version__)
        fig, ax = plt.subplots()
        ax.plot([1, 2, 3], [4, 5, 6])
        ax.set_title("Demo Plot")
        plt.close(fig)
        log.info("  Plot created and closed successfully.")
        check("Matplotlib test", True)
    except Exception as e:
        check(f"Matplotlib test failed: {e}", False)

    # --- Seaborn ---
    log.info("--- Seaborn ---")
    try:
        import seaborn as sns
        log.info("  Version: %s", sns.__version__)
        # Test a basic seaborn theme set and plot
        sns.set_theme(style="whitegrid")
        tips = pd.DataFrame({
            'total_bill': [10.34, 21.01, 23.68, 24.59],
            'tip': [1.66, 3.50, 3.31, 3.61],
            'day': ['Sun', 'Sun', 'Mon', 'Mon']
        })
        fig, ax = plt.subplots()
        sns.scatterplot(data=tips, x='total_bill', y='tip', hue='day', ax=ax)
        plt.close(fig)
        log.info("  Scatterplot created successfully.")
        check("Seaborn test", True)
    except Exception as e:
        check(f"Seaborn test failed: {e}", False)

    # --- Plotly ---
    log.info("--- Plotly ---")
    try:
        import plotly
        import plotly.express as px
        import plotly.graph_objects as go
        log.info("  Version: %s", plotly.__version__)
        fig = px.bar(x=["A", "B", "C"], y=[1, 2, 3], title="Test")
        html = fig.to_html(full_html=False, include_plotlyjs=False)
        assert len(html) > 0
        log.info("  Bar chart HTML generated (%d chars).", len(html))
        # Test graph_objects
        fig2 = go.Figure(data=[go.Indicator(mode="number", value=42)])
        assert fig2 is not None
        log.info("  Indicator figure created.")
        check("Plotly test", True)
    except Exception as e:
        check(f"Plotly test failed: {e}", False)

    # --- Folium ---
    log.info("--- Folium ---")
    try:
        import folium
        from folium.plugins import HeatMap
        log.info("  Version: %s", folium.__version__)
        m = folium.Map(location=[40.7128, -74.0060], zoom_start=10)
        heat_data = [[40.7128, -74.0060, 1.0]]
        HeatMap(heat_data).add_to(m)
        html = m._repr_html_()
        assert len(html) > 0
        log.info("  Map with HeatMap generated (%d chars).", len(html))
        check("Folium test", True)
    except Exception as e:
        check(f"Folium test failed: {e}", False)

    # --- FPDF2 ---
    log.info("--- FPDF2 ---")
    try:
        from fpdf import FPDF
        log.info("  Module: fpdf (fpdf2)")
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", size=12)
        pdf.cell(200, 10, text="Test PDF", align="C")
        # Output to bytes (don't write to disk)
        output = pdf.output()
        assert len(output) > 0
        log.info("  PDF generated in memory (%d bytes).", len(output))
        check("FPDF2 test", True)
    except Exception as e:
        check(f"FPDF2 test failed: {e}", False)

    # --- Openpyxl ---
    log.info("--- Openpyxl ---")
    try:
        import openpyxl
        log.info("  Version: %s", openpyxl.__version__)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Test'
        assert ws['A1'].value == 'Test'
        log.info("  Workbook created and cell written.")
        check("Openpyxl test", True)
    except Exception as e:
        check(f"Openpyxl test failed: {e}", False)

    # --- PyArrow ---
    log.info("--- PyArrow ---")
    try:
        import pyarrow as pa
        log.info("  Version: %s", pa.__version__)
        table = pa.table({'col1': [1, 2, 3], 'col2': ['a', 'b', 'c']})
        log.info("  Arrow table created with %d rows.", len(table))
        check("PyArrow test", True)
    except Exception as e:
        check(f"PyArrow test failed: {e}", False)

    # --- XlsxWriter ---
    log.info("--- XlsxWriter ---")
    try:
        import xlsxwriter
        log.info("  Version: %s", xlsxwriter.__version__)
        temp_path = os.path.join(os.path.dirname(__file__) or '.', '_test_temp.xlsx')
        wb = xlsxwriter.Workbook(temp_path)
        ws = wb.add_worksheet()
        ws.write('A1', 'Test')
        wb.close()
        assert os.path.exists(temp_path)
        os.remove(temp_path)
        log.info("  Workbook created, written, and cleaned up.")
        check("XlsxWriter test", True)
    except Exception as e:
        check(f"XlsxWriter test failed: {e}", False)
        temp_path = os.path.join(os.path.dirname(__file__) or '.', '_test_temp.xlsx')
        if os.path.exists(temp_path):
            os.remove(temp_path)

    # --- Streamlit ---
    log.info("--- Streamlit ---")
    try:
        import streamlit as st
        log.info("  Version: %s", st.__version__)
        # Minimal API check (no server run)
        assert hasattr(st, "set_page_config") and callable(st.set_page_config)
        assert hasattr(st, "write") and callable(st.write)
        log.info("  Run: streamlit run main_menu.py (from streamlit_demo/) to launch the demo app.")
        check("Streamlit test", True)
    except Exception as e:
        check(f"Streamlit test failed: {e}", False)

    # --- os (built-in) ---
    log.info("--- os (built-in) ---")
    try:
        log.info("  CWD: %s", os.getcwd())
        log.info("  Platform: %s", sys.platform)
        check("os test", True)
    except Exception as e:
        check(f"os test failed: {e}", False)

    # --- Summary ---
    log.info("=" * 55)
    log.info("  All library tests completed.")
    log.info("=" * 55)


if __name__ == "__main__":
    main()
