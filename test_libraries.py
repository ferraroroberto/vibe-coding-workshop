# Simple script to test basic imports and functionality of workshop libraries
# Testing libraries from requirements.txt in root folder
import os
import sys


def check(msg, success):
	# Green check for success, red cross for failure
	if success:
		print(f"\033[92m✔\033[0m {msg}")
	else:
		print(f"\033[91m✘\033[0m {msg}")


print("=" * 55)
print("  Workshop Library Test Suite")
print("=" * 55)
print(f"  Python: {sys.version.split()[0]}")
print(f"  CWD:    {os.getcwd()}")
print("=" * 55)
print()


# --- Faker ---
print("--- Faker ---")
try:
	import faker
	fake = faker.Faker()
	version = getattr(faker, '__version__', 'installed')
	print(f"  Version: {version}")
	print(f"  Sample name: {fake.name()}")
	check("Faker test", True)
except Exception as e:
	check(f"Faker test failed: {e}", False)
print()


# --- DuckDB ---
print("--- DuckDB ---")
try:
	import duckdb
	conn = duckdb.connect()
	result = conn.execute("SELECT 1 as test").fetchone()
	print(f"  Version: {duckdb.__version__}")
	print(f"  Test query result: {result}")
	conn.close()
	check("DuckDB test", True)
except Exception as e:
	check(f"DuckDB test failed: {e}", False)
print()


# --- Pandas ---
print("--- Pandas ---")
try:
	import pandas as pd
	print(f"  Version: {pd.__version__}")
	df = pd.DataFrame({'A': [1, 2], 'B': [3, 4]})
	assert len(df) == 2
	print(f"  DataFrame created: {df.shape}")
	check("Pandas test", True)
except Exception as e:
	check(f"Pandas test failed: {e}", False)
print()


# --- NumPy ---
print("--- NumPy ---")
try:
	import numpy as np
	print(f"  Version: {np.__version__}")
	arr = np.array([1, 2, 3, 4, 5])
	print(f"  Array mean: {arr.mean()}")
	check("NumPy test", True)
except Exception as e:
	check(f"NumPy test failed: {e}", False)
print()


# --- Matplotlib ---
print("--- Matplotlib ---")
try:
	import matplotlib
	matplotlib.use("Agg")  # Non-interactive backend
	import matplotlib.pyplot as plt
	print(f"  Version: {matplotlib.__version__}")
	fig, ax = plt.subplots()
	ax.plot([1, 2, 3], [4, 5, 6])
	ax.set_title("Demo Plot")
	plt.close(fig)
	print("  Plot created and closed successfully.")
	check("Matplotlib test", True)
except Exception as e:
	check(f"Matplotlib test failed: {e}", False)
print()


# --- Seaborn ---
print("--- Seaborn ---")
try:
	import seaborn as sns
	print(f"  Version: {sns.__version__}")
	# Test a basic seaborn theme set and plot
	sns.set_theme(style="whitegrid")
	tips = pd.DataFrame({
		'total_bill': [10.34, 21.01, 23.68, 24.59],
		'tip': [1.66, 3.50, 3.31, 3.61],
		'day': ['Sun', 'Sun', 'Mon', 'Mon']
	})
	fig, ax = plt.subplots()
	sns.scatterplot(data=tips, x='total_bill', y='tip', hue='day', ax=ax)
	plt.close(fig)
	print("  Scatterplot created successfully.")
	check("Seaborn test", True)
except Exception as e:
	check(f"Seaborn test failed: {e}", False)
print()


# --- Plotly ---
print("--- Plotly ---")
try:
	import plotly
	import plotly.express as px
	import plotly.graph_objects as go
	print(f"  Version: {plotly.__version__}")
	fig = px.bar(x=["A", "B", "C"], y=[1, 2, 3], title="Test")
	html = fig.to_html(full_html=False, include_plotlyjs=False)
	assert len(html) > 0
	print(f"  Bar chart HTML generated ({len(html)} chars).")
	# Test graph_objects
	fig2 = go.Figure(data=[go.Indicator(mode="number", value=42)])
	assert fig2 is not None
	print("  Indicator figure created.")
	check("Plotly test", True)
except Exception as e:
	check(f"Plotly test failed: {e}", False)
print()


# --- Folium ---
print("--- Folium ---")
try:
	import folium
	from folium.plugins import HeatMap
	print(f"  Version: {folium.__version__}")
	m = folium.Map(location=[40.7128, -74.0060], zoom_start=10)
	heat_data = [[40.7128, -74.0060, 1.0]]
	HeatMap(heat_data).add_to(m)
	html = m._repr_html_()
	assert len(html) > 0
	print(f"  Map with HeatMap generated ({len(html)} chars).")
	check("Folium test", True)
except Exception as e:
	check(f"Folium test failed: {e}", False)
print()


# --- FPDF2 ---
print("--- FPDF2 ---")
try:
	from fpdf import FPDF
	print(f"  Module: fpdf (fpdf2)")
	pdf = FPDF()
	pdf.add_page()
	pdf.set_font("Helvetica", size=12)
	pdf.cell(200, 10, text="Test PDF", align="C")
	# Output to bytes (don't write to disk)
	output = pdf.output()
	assert len(output) > 0
	print(f"  PDF generated in memory ({len(output)} bytes).")
	check("FPDF2 test", True)
except Exception as e:
	check(f"FPDF2 test failed: {e}", False)
print()


# --- Openpyxl ---
print("--- Openpyxl ---")
try:
	import openpyxl
	print(f"  Version: {openpyxl.__version__}")
	wb = openpyxl.Workbook()
	ws = wb.active
	ws['A1'] = 'Test'
	assert ws['A1'].value == 'Test'
	print("  Workbook created and cell written.")
	check("Openpyxl test", True)
except Exception as e:
	check(f"Openpyxl test failed: {e}", False)
print()


# --- PyArrow ---
print("--- PyArrow ---")
try:
	import pyarrow as pa
	print(f"  Version: {pa.__version__}")
	table = pa.table({'col1': [1, 2, 3], 'col2': ['a', 'b', 'c']})
	print(f"  Arrow table created with {len(table)} rows.")
	check("PyArrow test", True)
except Exception as e:
	check(f"PyArrow test failed: {e}", False)
print()


# --- XlsxWriter ---
print("--- XlsxWriter ---")
try:
	import xlsxwriter
	print(f"  Version: {xlsxwriter.__version__}")
	temp_path = os.path.join(os.path.dirname(__file__) or '.', '_test_temp.xlsx')
	wb = xlsxwriter.Workbook(temp_path)
	ws = wb.add_worksheet()
	ws.write('A1', 'Test')
	wb.close()
	assert os.path.exists(temp_path)
	os.remove(temp_path)
	print("  Workbook created, written, and cleaned up.")
	check("XlsxWriter test", True)
except Exception as e:
	check(f"XlsxWriter test failed: {e}", False)
	temp_path = os.path.join(os.path.dirname(__file__) or '.', '_test_temp.xlsx')
	if os.path.exists(temp_path):
		os.remove(temp_path)
print()


# --- os (built-in) ---
print("--- os (built-in) ---")
try:
	print(f"  CWD: {os.getcwd()}")
	print(f"  Platform: {sys.platform}")
	check("os test", True)
except Exception as e:
	check(f"os test failed: {e}", False)
print()


# --- Summary ---
print("=" * 55)
print("  All library tests completed.")
print("=" * 55)
